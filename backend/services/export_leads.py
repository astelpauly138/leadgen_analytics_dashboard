from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse
from supabase_client import supabase
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from typing import List

router = APIRouter()


def get_approved_leads(user_id: str):
    result = supabase.schema("analytics") \
        .from_("lead_analytics") \
        .select("*") \
        .eq("user_id", user_id) \
        .eq("status", "approved") \
        .execute()
    return result.data

def get_specific_leads_by_id(user_id: str, lead_ids: List[str]):
    result = supabase.schema("analytics").from_("lead_analytics") \
        .select("*") \
        .eq("user_id", user_id) \
        .in_("lead_id", lead_ids) \
        .execute()
    return result.data

def create_excel(data: list, file_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Data"

    if not data:
        return None

    headers = list(data[0].keys())
    ws.append(headers)

    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for row in data:
        ws.append(list(row.values()))

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(file_name)
    return file_name


@router.get("/export_specific_leads/{user_id}")
def export_specific_leads(user_id: str, lead_ids: List[str] = Query(None)):
    if not lead_ids:
        raise HTTPException(status_code=400, detail="No lead IDs provided")

    data = get_specific_leads_by_id(user_id, lead_ids)

    if not data:
        raise HTTPException(status_code=404, detail="No leads found for the selected IDs")

    file_name = "leads_data.xlsx"
    create_excel(data, file_name)

    return FileResponse(
        file_name,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=file_name
    )


@router.get("/export_approved/{user_id}")
def export_approved_leads(user_id: str):
    data = get_approved_leads(user_id)
    if not data:
        raise HTTPException(status_code=404, detail="No approved leads found")

    file_name = "leads_data.xlsx"
    create_excel(data, file_name)

    return FileResponse(
        file_name,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=file_name
    )
